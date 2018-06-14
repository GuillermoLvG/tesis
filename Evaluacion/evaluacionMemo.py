# -*- coding: utf-8 -*-
# Autor: Guillermo López Velarde González
#------------------------------------------------------
# No recibe argumentos.
# Devuelve métricas de evaluación. freelingNER.py vs main.py
#------------------------------------------------------
import os
import re
import openpyxl
import xml.etree.ElementTree
from pymongo import MongoClient
import csv
try:
	os.remove("tabla.csv")
	os.remove("tablaFL.csv")
	os.remove("tablaMemo.csv")
except OSError:
	pass

#Gold Standard
doc = openpyxl.load_workbook('entities_ajeno.xlsx')
sheet = doc['Sheet1']
entidades = []
row_count = sheet.max_row + 1
#desde empieza hasta 1 más
for reng in range(2,row_count):
	print (sheet.cell(row=reng,column=3).value)
	entidades.append([sheet.cell(row=reng,column=3).value,sheet.cell(row=reng,column=4).value])
entidades = [list(item) for item in set(tuple(row) for row in entidades)]
for entidad in entidades:
	escribir = csv.writer(open('tabla.csv', 'a+'))
	escribir.writerow([entidad[0],entidad[1]])

goldStandard = []
for entidad in entidades:
	valor = entidad[0]
	valor = valor.strip()
	valor = valor.replace(",","")
	valor = valor.replace(")","")
	valor = valor.replace("(","")
	valor = valor.replace("”","")
	valor = valor.replace("“","")
	valor = valor.replace(". ",".")
	goldStandard.append(valor.lower())
print (goldStandard)

#Resultado Freeling
client = MongoClient('localhost', 27017)
db = client.NERLegalesFL
collection = db.collection
result = collection.find({})
freelingResults = []
for element in result:
	valor = element["Nombre"]
	valor = valor.replace(" de el "," del ")
	valor = valor.replace(",","")
	valor = valor.replace(")","")
	valor = valor.replace("(","")
	valor = valor.replace(". ",".")
	freelingResults.append(valor.lower())
	escribir = csv.writer(open('tablaFL.csv', 'a+'))
	escribir.writerow([element["Nombre"],element["Clase"]])
#Calculo de precisión y exhaustividad
TP = 0
FP = 0
for element in freelingResults:
	if element in goldStandard:
		TP = TP + 1
	else:
		FP = FP + 1
precision = (TP/(TP+FP))*100
FN = 0
for element in goldStandard:
	if element not in freelingResults:
		FN = FN + 1
recall = (TP/(TP+FN))*100

print("Freeling")
print("Precision: " + str(precision))
print("Recall: " + str(recall))


#Resultado Memo
client = MongoClient('localhost', 27017)
db = client.NERLegales
collection = db.Entidades
result = collection.find({})
memoResults = []
for element in result:
	valor = element["Nombre"]
	valor = valor.replace(" de el "," del ")
	valor = valor.replace(". ",".")
	memoResults.append(valor.lower())
	escribir = csv.writer(open('tablaMemo.csv', 'a+'))
	escribir.writerow([element["Nombre"],element["Clase"]])

#Calculo de precisión y exhaustividad
TP = 0
FP = 0
for element in memoResults:
	if element in goldStandard:
		#print ("Verdadero Positivo:" + element)
		TP = TP + 1
	else:
		FP = FP + 1
		print("Falso Positivo:" + element)
precision = (TP/(TP+FP))*100
FN = 0
for element in goldStandard:
	if element not in memoResults:
		FN = FN + 1
		print("Falso Negativo:" + element)
recall = (TP/(TP+FN))*100

print("Memo")
print("Precision:" + str(precision))
print("Recall: " + str(recall))
