# -*- coding: utf-8 -*-
# Autor: Guillermo López Velarde González
#------------------------------------------------------
# No recibe argumentos.
# Devuelve métricas de evaluación. freelingNER.py vs main.py
#------------------------------------------------------
import os
import re
import openpyxl
import xml.etree.ElementTree
from pymongo import MongoClient

#xlsx de Documentos
doc = openpyxl.load_workbook('ResultadosEtiquetados/documents.csv.xlsx')
sheet = doc['documents.csv']
documentos = dict()
for reng in range(2,71):
	documentos[reng-1] = sheet.cell(row=reng,column=2).value.replace("../DOCX/","")

#xlsx de Ocurrencias
doc = openpyxl.load_workbook('ResultadosEtiquetados/occurences.csv.xlsx')
sheet = doc['occurences.csv']
occurences = dict()
for doc_id in range(1,70):
	occurences[doc_id] = set()
	for reng in range(2,11205):
		document_id_xlsx = sheet.cell(row=reng,column=2).value
		entity_type = sheet.cell(row=reng,column=4).value
		entity_id = sheet.cell(row=reng,column=3).value
		if doc_id == document_id_xlsx and entity_type == "entities":
			occurences[doc_id].add(entity_id)

#xlsx de Entidades
doc = openpyxl.load_workbook('ResultadosEtiquetados/entities.csv.xlsx')
sheet = doc['entities.csv']
entities = dict()
for reng in range(1,1551):
	entidad_id = sheet.cell(row=reng,column=1).value
	entities[entidad_id] = sheet.cell(row=reng,column=2).value

#Fusión de Ocurrencias y Entidades
occurencesEntities = dict()
for key,value in occurences.items():
	occurencesEntities[key] = []
	listaEntidades = list(value)
	for entidad_id in listaEntidades:
		try:
			occurencesEntities[key].append(entities[entidad_id])
		except KeyError as e:
			occurencesEntities[key].append(entidad_id)

#Gold Standard			
goldStandard = []
for entidad in occurencesEntities[26]: #Cambiar el número para cambiar de archivo a evaluar
	if type(entidad) == str:
		goldStandard.append(entidad.lower())

#Resultado Freeling
client = MongoClient('localhost', 27017)
db = client.NERLegalesFL
collection = db.collection
result = collection.find({})
freelingResults = []
for element in result:
	freelingResults.append(element["Nombre"].replace("_"," ").lower())

#Calculo de precisión y exhaustividad
TP = 0
FP = 0
for element in freelingResults:
	if element in goldStandard:
		TP = TP + 1
	else:
		FP = FP + 1
precision = (TP/(TP+FP))*100
FN = 0
for element in goldStandard:
	if element not in freelingResults:
		FN = FN + 1
recall = (TP/(TP+FN))*100

print("Freeling")
print("Precision: " + str(precision))
print("Recall: " + str(recall))

#Resultado Memo
client = MongoClient('localhost', 27017)
db = client.NERLegales
collection = db.Entidades
result = collection.find({})
memoResults = []
for element in result:
	memoResults.append(element["Nombre"].lower())

#Calculo de precisión y exhaustividad
TP = 0
FP = 0
for element in memoResults:
	if element in goldStandard:
		TP = TP + 1
	else:
		FP = FP + 1
precision = (TP/(TP+FP))*100
FN = 0
for element in goldStandard:
	if element not in memoResults:
		FN = FN + 1
recall = (TP/(TP+FN))*100

print("Memo")
print("Precision:" + str(precision))
print("Recall: " + str(recall))
